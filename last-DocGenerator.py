#-*- coding: utf-8 -*-
'''

'''
import os
import re
import corpinfo #自制企业信息模块
import chntoday #自制当天年月日格式
from docxtpl import DocxTemplate, InlineImage #根据DOCX模板生成结果用
from docx.shared import Mm, Inches, Pt  #图像 for height and width you have to use millimeters (Mm), inches or points(Pt) class :
from openpyxl import Workbook  #读取XLSX文件用
from openpyxl import load_workbook  #读取XLSX文件用
from openpyxl.utils import get_column_letter #读取XLSX文件用

class corp: #固定的企业信息，从内部查询
    def __init__(self):
        self.corpname = ""
        self.addr = ""
        self.regnum = ""
        self.phone = ""
        self.repperson = ""
        self.date = ""
        self.calldate = ""
        self.callhour = ""
        self.callmin = ""
        self.imgexp = ""
        self.recexp = ""
        self.marker = " "
        self.corpindex = " "
        
        self.index = 0 #初始化
        self.rootdir = os.getcwd()
        self.successdir = []
        self.faildir = []
        self.found = 0 #核查表是否匹配到企业的标识
        self.cwd = ''
        
        
    def processallfolders(self):
       for singledir,subdirs,files in os.walk(self.rootdir):
           #print(singledir)
           #print("")
           if('lib' not in singledir):
               if(self.processfolder(singledir)): #确定内置企业数据库中有没有目标企业
                   print('正在处理：' + self.corpname)
                   print('......')
                   self.cwd = singledir + '\\'
                   if(self.inspectrecord()): #确定核查表中有没有对应企业的核查信息
                       self.generateImgDoc()
                       self.generateInspectRecord()
                       print('成功生成文书！')
                       print('****************')
                       print(' ')
                       self.successdir.append(singledir)
                   elif('lib' not in singledir):
                       self.faildir.append(singledir)                
               elif('lib' not in singledir):
                   self.faildir.append(singledir)            

    def processfolder(self,folder_path):
                
        self.corpname = re.sub(r'.*-',"",folder_path) #删除剩下企业名称
        self.corpname = re.sub(r'.*\\',"",self.corpname) #删除剩下企业名称
        print('尝试匹配文件夹"' + self.corpname + '"')
        print('......')
        #以下根据企业名称查询信息
        try:
            self.addr = corpinfo.allcorpinfo[self.corpname]['addr']
            if(corpinfo.phone[self.corpname]):
                self.phone = corpinfo.phone[self.corpname]
            self.regnum = corpinfo.allcorpinfo[self.corpname]['regnum'] 
            self.repperson = corpinfo.allcorpinfo[self.corpname]['repperson']
            self.date = chntoday.chntoday 
            self.explanation = '现场无该企业经营的迹象，该企业通过登记地址无法联系。'
            return True
        except:
            #print("数据库中无此企业：%s" % self.corpname)
            print("数据库中无此企业")
            print('****************')
            print(" ")
            return False
            
    def inspectrecord(self): #从外部读取的检查情况、日期等信息
        cwdfiles = os.listdir(self.rootdir) #列出py文件所在目录的文件
        for file in cwdfiles:               #逐个检测文件名是否符合要求
            #print(file)
            if(file.lower() == "企业信息及核查记录表.xlsx"):  #如果找到符合的，就更改wbpath
                wbpath = self.rootdir + "\\" + file
                #print(wbpath)
        try:                    #尝试读取wbpath，如果上一步没有找到文件，则会提示没有找到文件的错误
            self.found = 0
            wb = load_workbook(filename=wbpath)  
            sheets = wb.get_sheet_names()
            #第一个表格的名称
            sheet_first = sheets[0]
            #获取特定的worksheet
            ws = wb.get_sheet_by_name(sheet_first)
            rows = ws.rows
            column = ws.columns
            for row in rows:
                if (row[2].value == self.corpname):
                    self.marker = row[0].value
                    self.corpindex = row[1].value
                    
                    self.date = row[9].value
                    self.hourmin = row[10].value
                    self.endhourmin = row[11].value
                    self.imgexp = str(row[12].value)
                    self.recexp = str(row[12].value)
                    self.calldate = row[13].value
                    self.callhour = row[14].value
                    self.callmin = row[15].value
                    self.callresult = row[16].value
                    self.askingphoto = row[17].value

                    self.found = 1
                    break
            if self.found <1:
                print('在记录表中没有找到'+ self.corpname +'的核查记录，请检查')
                return False
            else:
                return True                
        except:
            #print(cwdfiles)
            print("找不到名为'企业信息及核查记录表.xlsx'的文件，请确认与本文件放在同一文件夹中。")
            os.system('pause')
                
    def generateImgDoc(self): #生成证据单的函数，用docxtpl模块，以tpl指定的文件为模板进行元素替换。
        self.index = 0
        for file in os.listdir(self.cwd): #历遍图片文件
            if ('jpg' in file.lower() or 'jpeg'in file.lower()): #判断文件名是否图片
                self.index = self.index + 1 #找到后，序号加1
                try:
                    tpl=DocxTemplate(self.rootdir + '\\证据提取单模板.docx') #指定的模板
                    image = self.cwd + '\\' + file
                    ImgDocPath = self.cwd + '\\' + self.corpname + '-照片-'+ str(self.index) + '.docx' #路径（全局变量）+字号+序号+格式
                    context = {
                        'image' : InlineImage(tpl,image,width=Mm(153)), #替换图片
                        'date' : self.date,  #替换日期
                        'explanation' : '以上为执法人员于' + str(self.date) + '对位于' + str(self.addr) + '的' + self.corpname + '进行核查时的照片。' + self.imgexp,  #替换说明
                        'marker': self.marker,
                        'corpindex' : self.corpindex
                    }
                    tpl.render(context) #执行替换
                    tpl.save(ImgDocPath) #保存文件
                except:
                    print('当前目录未找到“证据提取单模板.docx”，请确认文件已经放入本文件夹且文件名正确')
                    os.system('pause')
                    exit(0)

        
    def generateInspectRecord(self): #生成现场笔录的函数
        try:
            tpl=DocxTemplate(self.rootdir +'\\现场笔录模板.docx') #指定的模板
                    #确定文件保存路径
            RecordPath = self.cwd + '\\' + self.corpname + '-现场笔录-' + '.docx' #路径（全局变量）+字号+格式
            asking = ''
        ##如果表格中记录有询问照片，则加入相关表述。
            if(self.askingphoto == "是"):
                asking = '我执法人员通过问询周边业户得知，位于'+self.addr+'的'+ self.corpname + '，已不在此场所从事经营活动，去向未知。'
        	#确定替换的内容
            context = {
                'corpname' : self.corpname,
                'addr' : self.addr,
                'date' : self.date,
                'phone' : self.phone,
                'hourmin' : self.hourmin,
                'endhourmin' : self.endhourmin,
                'regnum' : self.regnum,
                'repperson' : self.repperson,
                'recexp' : self.recexp,
                'asking' : asking,
                'calldate' : self.calldate,
                'callhour' : self.callhour,
                'callmin' : self.callmin,
                'callresult' : self.callresult,
                'marker': self.marker,
                'corpindex' : self.corpindex                
            }
            tpl.render(context)
            tpl.save(RecordPath)
        except:
            print('当前目录未找到“现场笔录模板.docx”，请确认文件已经放入本文件夹且文件名正确')
            os.system('pause')
            exit(0)

        
    def excute(self): #执行全部步骤
        try:
            self.inspectrecord(self,self.currentfolder)
            self.generateImgDoc()
            self.generateInspectRecord()
        except:
            print("出现错误")
            
    def printresult(self): #打印结果
        print('-----------处理结果---------------------')
        print("成功在下列文件夹生成文书：")
        for item in range(len(corp.successdir)):
            print(str(item + 1) + ': ' + corp.successdir[item])
            print('')
        print("由于在核查记录表或企业信息库中未匹配到企业字号，下列文件夹未成功处理：")
        for item in range(len(corp.faildir)):
            print(str(item +1 ) + ': ' + corp.faildir[item])
        os.system('pause')
            

        
if __name__ == '__main__':
    corp = corp()
    corp.processallfolders()
    corp.printresult()

    
    
    
