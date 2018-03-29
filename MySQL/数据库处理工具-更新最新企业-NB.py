#-*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from string import Template
from datetime import date
import os


class data:
    def __init__(self):
        self.datacontent = ''
        self.datatpl = Template("('${c}','${r}','${a}','${rp}', '${cp}', 'active', '${div}'),\n")
        self.div = ''
        self.headtpl = ''
        self.head = ''
        now = date.today()
        self.today = "%d-%d-%d" %(now.year, now.month, now.day -1)

    def div_select(self):
        self.headtpl = Template('''
        UPDATE `hdscjg_database`.`ALL_corp` SET `Active` = 'not_active' WHERE `division` = '${div}';
        insert into `hdscjg_database`.`ALL_corp` 
        (`CorpName`, `RegNum`, `Addr`, `RepPerson`, `ContactPerson`, `Active`, `division` ) VALUES
        ''')
        print('请在下列名单中选择对应的监管所代码：\n'
              '1、SL 狮岭\n'
              '2、YH 裕华\n'
              '3、TB 炭步\n'
              '请输入两个英文字符代码(必须大写):')
        self.div = input()
        self.head = self.headtpl.substitute(div=self.div)

    def load_workbook(self):
        print('正在读取XLSX文件中的用户名单......')
        try:
            wb = load_workbook('更新最新企业.xlsx')
            self.ws = wb.worksheets[0]
        except FileNotFoundError:
            print('当前目录没有“更新最新企业.xlsx”文件')
            exit(0)

    def process_to_sql(self):
        print('正在处理用户名单......')
        rows = self.ws[3:self.ws.max_row]
        for row in rows:
            c = r = a = rp = cp = ''
            try:
                c = ''.join(row[0].value.split())
            except AttributeError:
                c = '(' + str(r) + ')无字号'

            try:
                r = ''.join(row[1].value.split())
            except AttributeError:
                print('存在注册号为空的情况，请检查后重新运行。')

            try:
                a = ''.join(row[2].value.split())
            except AttributeError:
                a = ''

            try:
                p = ''.join(row[3].value.split())
            except AttributeError:
                p = ''

            # 以下读取年报情况列，并根据年报情况处理导入的内容，对于未填报者，增加更新时间标识。
            try:
                nb = ''.join(row[4].value.split())
            except AttributeError:
                nb = row[4].value
            if nb == '未填报':
                nb = '17年度：截至' + self.today + nb
            else:
                nb = '17年度：' + nb

            try:
                rp = ''.join(row[7].value.split())
            except AttributeError:
                rp = ''

            try:
                cp = ''.join(row[8].value.split())
            except AttributeError:
                cp = ''

            try:
                cph = ''.join(row[9].value.split())
            except AttributeError:
                cph = ''

            try:
                ins = row[10].value 
            except IndexError:
                pass
            try:
                phcal = row[11].value
            except IndexError:
                pass
            assert c != ''
            self.datacontent += self.datatpl.substitute(c=c, r=r, a=a, rp=rp, cp=cp, div=self.div)

    def save_file(self):
        f = open(self.div + '-最新企业（无年报信息）.sql', 'wb')
        f.write(self.head.encode('utf8'))
        f.write(self.datacontent[:-2].encode('utf8'))
        f.write(b'''
        on duplicate key update 
        CorpName = Values(CorpName),
        Addr=values(addr),
        repperson = values(repperson),
        contactperson = values(contactperson),
        division = values(division);''')
        f.close()


if __name__ == '__main__':
    data = data()
    data.div_select()
    data.load_workbook()
    data.process_to_sql()
    data.save_file()