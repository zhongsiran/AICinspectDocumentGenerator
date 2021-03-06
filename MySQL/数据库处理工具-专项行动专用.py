#-*- coding: utf-8 -*-
from string import Template

from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet


class Data:
    def __init__(self):
        self.datacontent = ''
        self.datatpl = Template("('${sano}','${n}','${no}','${pre_r}', '${pre_n}', '${c}','${r}','${sd}','${ed}'),\n")
        self.div = ''
        self.ws = Worksheet
        self.head_tpl = ''
        self.head = ''

    def div_select(self):
        self.head_tpl = Template('''
insert into `hdscjg_database`.`${div}_zhuan_xiang_xing_dong`
(`sp_action_no`,`sp_action_name`, `sp_action_daihao`, `sp_action_pre_regnum`, `sp_action_pre_name`, `sp_action_corpname`, `sp_action_regnum`, `sp_action_startdate`, `sp_action_enddate`) VALUES
''')
        self.div = input()
        self.head = self.head_tpl.substitute(div=self.div)

    def load_workbook(self):
        try:
            wb = load_workbook('专项行动记录表.xlsx')
            self.ws = wb.worksheets[0]
        except FileNotFoundError:
            print('当前目录没有“专项行动记录表.xlsx”文件')
            exit(0)

    def process_to_sql(self):

        rows = self.ws[4:self.ws.max_row]  # 第一、二行是说明，第三行是标题
        for row in rows:
            try:
                r = ''.join(row[4].value.split())
            except AttributeError:
                print('存在注册号为空的情况，请检查后重新运行。')

            try:
                no = ''.join(row[0].value.split())
            except AttributeError:
                print('存在行动代号为空的情况，请检查后重新运行。')
                
            try:
                x = ''.join(row[1].value.split())
            except AttributeError:
                print('存在行动名为空的情况，请检查后重新运行。')
            try:
                daihao = ''.join(row[2].value.split())
            except AttributeError:
                daihao = row[2].value
            try:
                corpname = ''.join(row[3].value.split())
            except AttributeError:
                corpname = ''
            try:
                startdate = ''.join(row[5].value.split())
            except AttributeError:
                startdate = ''
            try:
                enddate = ''.join(row[6].value.split())
            except AttributeError:
                enddate = ''

            try:
                pn = ''.join(row[7].value.split())
            except AttributeError:
                pn = row[7].value

            try:
                pr = ''.join(row[8].value.split())
            except AttributeError:
                pr = row[8].value

            assert daihao != ''
            assert r or pr != ''

            self.datacontent += self.datatpl.substitute(sano=no, n=x, no=daihao, pre_r=pr, pre_n=pn, c=corpname, r=r, sd=startdate, ed=enddate)

    def savefile(self):
        f = open(self.div + '_zhuan_xiang.sql', 'wb')
        f.write(self.head.encode('utf8'))
        f.write(self.datacontent[:-2].encode('utf8'))
        f.close()


if __name__ == '__main__':
    data = Data()
    data.div_select()
    data.load_workbook()
    data.process_to_sql()
    data.savefile() 


