from openpyxl import load_workbook   # 读取XLSX文件用

wb_path = 'test.xlsx'
wb = load_workbook(wb_path)

attribute_dict = {}
ws =  wb[wb.sheetnames[0]]
the_row = ws.iter_rows(min_row=1, max_row=1)
for cells in the_row:
    for cell in cells:
        attribute_dict[cell.value] = cell.col_idx

print(attribute_dict)