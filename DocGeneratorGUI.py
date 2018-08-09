import os
import re
from docxtpl import DocxTemplate, InlineImage   # 根据DOCX模板生成结果用
from docx.shared import Mm    # 图像 for height and width you have to use millimeters (Mm), inches or points(Pt) class :
from openpyxl import load_workbook   # 读取XLSX文件用
import threading
from wx.lib.pubsub import pub
from PIL import Image

thread_lock = threading.Lock()


class DocGeneratorMainThread(threading.Thread):
    """docstring for doc_generator_main_thread"""
    def __init__(self, original_path, target_path, workbook_path, ins_tpl_path, img_tpl_path):
        super(DocGeneratorMainThread, self).__init__()
        self.setDaemon(True)
        self.running = True
        self.doc_generator_main = DocGenerator(original_path, target_path, workbook_path, ins_tpl_path, img_tpl_path)

    def run(self):
        self.doc_generator_main.process_all_folders()
        self.doc_generator_main.print_result()
        self.doc_generator_main.save_result()
        self.doc_generator_main.del_wb()


class DocGenerator:   # 固定的企业信息，从内部查询
    def __init__(self, original_path, target_path, workbook_path, ins_tpl_path, img_tpl_path):
        # 初始化企业信息
        self.corp_table_attributes = {'企业名称': 2}
        self.attributes_values = {'address': '', 'registration_num': '', 'phone': '', 'represent_person': '',
                                      'date': ''}
        self.corp_name = ""
        # self.address = ""
        # self.registration_num = ""
        # self.phone = ""
        # self.represent_person = ""
        # self.date = ""
        # self.call_date = ""
        # self.call_hour = ""
        # self.call_min = ""
        # self.image_explanation = ""
        # self.record_explanation = ""
        # self.marker = " "
        # self.corp_index = " "
        # self.hour_min = ''
        # self.end_hour_min = ''
        # self.call_result = ''
        # self.asking_photo = False

        # 初始化工作路径
        self.original_root_dir = original_path
        self.original_current_path = ''
        self.target_root_dir = target_path
        self.target_current_path = ''
        self.workbook_path = workbook_path
        self.ins_tpl_path = ins_tpl_path
        self.img_tpl_path = img_tpl_path

        # 初始化结果统计
        self.success_dirs = []
        self.failed_dirs = []

        self.load_ins_workbook(workbook_path)

    def process_all_folders(self):   # 主处理函数
        for singledir, subdirs, files in os.walk(self.original_root_dir):
            if 'lib' not in singledir:
                if self.corp_folder_match(singledir):   # 先从内置企业数据库中找、再从核查表中取得目标企业核查信息
                    self.original_current_path = singledir + '\\'
                    try:
                        os.mkdir(self.target_root_dir + '\\' + self.corp_name)
                    except FileExistsError:
                        pass
                    finally:
                        self.target_current_path = self.target_root_dir + '\\' + self.corp_name
                    
                    try:
                        self.generate_img_doc()
                        self.generate_inspect_record()
                        post_progress('成功生成文书！')
                        self.success_dirs.append(singledir)
                    except Exception:
                        post_progress('生成过程中出错')
                        self.failed_dirs.append(singledir)
                elif 'lib' not in singledir:
                    self.failed_dirs.append(singledir)
            elif'lib' not in singledir:
                self.failed_dirs.append(singledir)

    def corp_folder_match(self, folder_path):
        self.corp_name = re.sub(r'.*-', "", folder_path)   # 删除剩下企业名称
        self.corp_name = re.sub(r'.*\\', "", self.corp_name)   # 删除剩下企业名称
        post_progress('尝试匹配文件夹"' + self.corp_name + '"')

        # 以下根据企业名称查询信息
        if self.get_corp_inspect_record():   # 尝试从核查表读取
            return True
        else:
            return False

    def load_ins_workbook(self, workbook_path):
        try:   # 尝试读取wbpath
            wb = load_workbook(filename=workbook_path)  
            self.ws = wb[wb.sheetnames[0]]  # 打开全局性的纪录表
        except FileNotFoundError:
            post_progress('指定了不存在的核查表文件')
            post_finished('无法打开核查表，请检查是否选择错误')
            raise FileNotFoundError

    def get_corp_inspect_record(self):  # 从外部读取的检查情况、日期等信息
        if not self.ws:
            self.load_ins_workbook(self.workbook_path)
        elif self.corp_table_attributes == {'企业名称': 2}:
            first_row = self.ws.iter_rows(min_row=1, max_row=1)
            for cells in first_row:
                for cell in cells:
                    if cell.value:
                        self.corp_table_attributes[cell.value] = cell.col_idx - 1
        else:
            # print(self.corp_table_attributes)
            found = False
            rows = self.ws.iter_rows(min_row=2, max_row=self.ws.max_row)
            for row in rows:
                print('企业名称对应的列是：' + str(self.corp_table_attributes['企业名称'] + 1))
                if row[self.corp_table_attributes['企业名称']].value == self.corp_name:
                    for attribute_name in self.corp_table_attributes:
                        if row[self.corp_table_attributes[attribute_name]].value != '':  # 遍历取得参数
                            self.attributes_values[attribute_name] = row[self.corp_table_attributes[attribute_name]].value
                        else:
                            self.attributes_values[attribute_name] = ''

                    found = True   # 表示成功取得本户应有资料
                    break
                else:
                    found = False   # 再次赋值作为提示

            if found is False:
                # post_progress('在记录表中没有找到'+ self.corpname +'的核查记录')
                return False   # 后接corp_folder_match
            else:
                return True   # 后接corp_folder_match
                
    def generate_img_doc(self):   # 生成证据单的函数，用docxtpl模块，以tpl指定的文件为模板进行元素替换。
        index = 0   # 图片编号
        for file in os.listdir(self.original_current_path):   # 历遍图片文件
            if 'jpg' in file.lower() or 'jpeg'in file.lower():   # 判断文件名是否图片
                index = index + 1   # 找到后，序号加1
                try:
                    tpl = DocxTemplate(self.img_tpl_path)  # 指定的模板
                    image = self.original_current_path + file
                    img_doc_path = self.target_current_path + '\\' + self.corp_name + '-照片-' + str(index) + '.docx'  # 路径（全局变量）+字号+序号+格式

                    img = Image.open(image)
                    if img.size[0] > img.size[1]:
                        self.attributes_values['image'] = InlineImage(tpl, image, width=Mm(153))  # 替换图片
                    else:
                        self.attributes_values['image'] = InlineImage(tpl, image, width=Mm(130))  # 替换图片

                    try:
                        tpl.render(self.attributes_values)  # 执行替换
                        tpl.save(img_doc_path)  # 保存文件
                    except UnrecognizedImageError:
                        content = file + '不是有效的图片文件，无法生成证据提取单'
                        post_progress(content)
                except Exception:
                    self.failed_dirs.append('处理' + file + '的图片时出错')
            else:
                pass

    def generate_inspect_record(self):  # 生成现场笔录的函数
        try:
            tpl = DocxTemplate(self.ins_tpl_path)   # 指定的模板F
            # 确定文件保存路径
            record_path = self.target_current_path + '\\' + self.corp_name + '-现场笔录' + '.docx'   # 路径（全局变量）+字号+格式
            # 确定替换的内容
            tpl.render(self.attributes_values)
            tpl.save(record_path)
        except Exception as e:
            print('210' + e)
            
    def print_result(self):   # 打印结果
        post_progress('####' * 15)
        post_progress('''
核查文书生成器处理结果
制作单位：狮岭监管所
联系人：钟思燃
联系电话：661668
--------处理结果---------------------''')
        if len(self.success_dirs) > 0:
            post_progress("成功在下列文件夹生成文书：")
            for item in range(len(self.success_dirs)):
                post_progress(str(item + 1) + ': ' + self.success_dirs[item])
                post_progress('')
        post_progress('更多详情请看“处理结果.txt文件”')
        post_finished('处理完毕')

    def save_result(self):
        result_file = open(self.target_root_dir + '\\' + '处理结果.txt','w+')
        result_file.write('''
核查文书生成器处理结果
制作单位：狮岭监管所
联系人：钟思燃
联系电话：661668

-------------------处理结果---------------------\n''')
        if len(self.success_dirs) > 0:
            result_file.write("成功在下列文件夹生成文书：\n")
            for item in range(len(self.success_dirs)):
                result_file.write(str(item + 1) + ': ' + self.success_dirs[item] + '\n')
                result_file.write('-----------------------------------------------\n')
        if len(self.failed_dirs) > 0 :
            result_file.write("由于在核查记录表或企业信息库中未匹配到企业字号等原因，下列文件夹或文件未成功处理：\n")
            for item in range(len(self.failed_dirs)):
                result_file.write(str(item + 1) + ': ' + self.failed_dirs[item] + '\n')

    def del_wb(self):
        del self.ws


def post_progress(msg_to_post):
    pub.sendMessage("update_dg", msg=msg_to_post)


def post_finished(msg_to_post):
    pub.sendMessage("dg_finished", result=msg_to_post)

